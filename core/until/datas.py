# -*- coding: utf-8 -*-
""" 
@Filename:  core/until/datas
@Author:    小蔡 
@Time:      2023/2/24 11:17
@Describe: ... 
"""

from openpyxl import Workbook, load_workbook


def load_case_by_excel(file):
    """从excel文件中加载用例"""
    wb: Workbook = load_workbook(file)  # 打开文件

    suite_list = []

    for ws in wb.worksheets:

        case_list = []  # 存放用例的列表
        suite = {
            "name": ws.title,  # 工作表的名字
            "case_list": case_list,  # 放入字典
        }

        suite_list.append(suite)

        # 如果让同名的旧变量，活下来

        for line in ws.iter_rows(values_only=True, min_row=2):  # 从工作表中逐行读取  值

            # 如果第一列的值是0，说明进入了新的测试用例， 关键字参数，就是用例的名称
            if line[0] == 0:
                # print("进入了新的用例，他的名字是", line[3])
                case = {  # 创建新用例
                    "name": "",  # 用例名称
                    "steps": [],  # 用例步骤
                }
                case_list.append(case)
                case["name"] = line[3]
            else:
                # print("    测试步骤", line)
                case["steps"].append(line)
                # 如果第一列的值不是0，说明这一行值，是测试用例的步骤

    return suite_list
